"""Profile-related background tasks — Excel import, certificate generation, photo resize."""

import base64
import io
import json
from uuid import UUID

import structlog
from PIL import Image

from app.tasks import broker

logger = structlog.get_logger(__name__)

MAX_PHOTO_SIZE = (800, 800)


@broker.task  # type: ignore[misc]
async def process_excel_import(task_id: str, file_bytes_b64: str) -> None:
    """Decode base64 Excel file, parse rows and create doctor profiles.

    Progress is written to Redis key ``import:{task_id}``.
    """
    from datetime import UTC, datetime
    from io import BytesIO
    from typing import Any
    from uuid import uuid4

    from openpyxl import load_workbook
    from redis.asyncio import Redis
    from sqlalchemy import select

    from app.core.database import AsyncSessionLocal
    from app.core.redis import get_redis_pool
    from app.core.security import hash_password
    from app.models.profiles import DoctorProfile
    from app.models.users import User

    file_bytes = base64.b64decode(file_bytes_b64)
    redis: Redis = Redis(connection_pool=get_redis_pool())  # type: ignore[type-arg]

    errors: list[dict[str, Any]] = []
    imported = 0
    total_rows = 0

    async with AsyncSessionLocal() as db:
        try:
            wb = load_workbook(filename=BytesIO(file_bytes), read_only=True)
            ws = wb.active
            if ws is None:
                await redis.set(
                    f"import:{task_id}",
                    json.dumps({
                        "status": "error",
                        "total_rows": 0,
                        "imported": 0,
                        "errors": [{"row": 0, "error": "Empty workbook"}],
                    }),
                    ex=3600,
                )
                return

            rows = list(ws.iter_rows(min_row=2, values_only=True))
            total_rows = len(rows)

            for idx, row in enumerate(rows, start=2):
                try:
                    email = str(row[0]).strip() if row[0] else ""
                    first_name = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                    last_name = str(row[2]).strip() if len(row) > 2 and row[2] else ""
                    phone = str(row[3]).strip() if len(row) > 3 and row[3] else ""

                    if not email:
                        errors.append({"row": idx, "error": "Email is required"})
                        continue

                    existing = await db.execute(select(User.id).where(User.email == email))
                    if existing.scalar_one_or_none():
                        errors.append({"row": idx, "error": f"Duplicate email: {email}"})
                        continue

                    temp_password = f"Tmp{uuid4().hex[:12]}!"
                    user = User(
                        email=email,
                        password_hash=hash_password(temp_password),
                        email_verified_at=datetime.now(UTC),
                        is_active=True,
                    )
                    db.add(user)
                    await db.flush()

                    profile = DoctorProfile(
                        user_id=user.id,
                        first_name=first_name or "N/A",
                        last_name=last_name or "N/A",
                        phone=phone or "",
                        status="approved",
                    )
                    db.add(profile)
                    await db.flush()
                    imported += 1

                except Exception as exc:
                    errors.append({"row": idx, "error": str(exc)})

            await db.commit()
            status = "completed"
        except Exception as exc:
            await db.rollback()
            logger.error("import_failed", error=str(exc))
            errors.append({"row": 0, "error": f"File processing error: {exc}"})
            imported = 0
            status = "error"

    result_data = {
        "status": status,
        "total_rows": total_rows,
        "imported": imported,
        "errors": errors,
    }
    await redis.set(f"import:{task_id}", json.dumps(result_data), ex=3600)
    logger.info("excel_import_done", task_id=task_id, imported=imported, errors=len(errors))


@broker.task  # type: ignore[misc]
async def generate_certificate(doctor_profile_id: str) -> None:
    """Generate a membership certificate PDF for a doctor profile, idempotently."""
    from datetime import UTC, datetime

    from sqlalchemy import select

    from app.core.database import AsyncSessionLocal
    from app.models.certificates import Certificate
    from app.services.certificate_service import CertificateService

    try:
        async with AsyncSessionLocal() as db:
            current_year = datetime.now(UTC).year
            existing = (
                await db.execute(
                    select(Certificate.id).where(
                        Certificate.doctor_profile_id == UUID(doctor_profile_id),
                        Certificate.certificate_type == "member",
                        Certificate.year == current_year,
                        Certificate.is_active.is_(True),
                    )
                )
            ).scalar_one_or_none()
            if existing:
                logger.info(
                    "certificate_already_exists",
                    profile_id=doctor_profile_id, year=current_year,
                )
                return

            svc = CertificateService(db)
            await svc.generate_membership_certificate(UUID(doctor_profile_id))
            logger.info("certificate_generated_task", profile_id=doctor_profile_id)
    except Exception:
        logger.exception("generate_certificate_failed", profile_id=doctor_profile_id)


@broker.task  # type: ignore[misc]
async def resize_profile_photo(s3_key: str) -> None:
    """Download a profile photo from S3, resize to 800x800, re-upload."""
    try:
        from app.services import file_service

        session = file_service._get_s3_session()
        async with session.client(**file_service._s3_client_kwargs()) as s3:
            response = await s3.get_object(
                Bucket=file_service.settings.S3_BUCKET,
                Key=s3_key,
            )
            data = await response["Body"].read()

        img = Image.open(io.BytesIO(data))
        img = img.convert("RGB")
        img.thumbnail(MAX_PHOTO_SIZE, Image.Resampling.LANCZOS)

        buf = io.BytesIO()
        img.save(buf, format="JPEG", quality=85)
        resized_bytes = buf.getvalue()

        async with session.client(**file_service._s3_client_kwargs()) as s3:
            await s3.put_object(
                Bucket=file_service.settings.S3_BUCKET,
                Key=s3_key,
                Body=resized_bytes,
                ContentType="image/jpeg",
            )

        logger.info("photo_resized", s3_key=s3_key)
    except Exception:
        logger.exception("resize_profile_photo_failed", s3_key=s3_key)
