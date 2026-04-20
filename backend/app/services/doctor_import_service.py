"""Doctor import from Excel files."""

from __future__ import annotations

import json
from datetime import UTC, datetime
from io import BytesIO
from typing import Any
from uuid import uuid4

import structlog
from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.enums import DoctorStatus
from app.core.exceptions import NotFoundError
from app.core.security import hash_password
from app.core.utils import generate_unique_slug
from app.models.profiles import DoctorProfile
from app.models.users import User
from app.schemas.doctor_admin import ImportErrorItem, ImportStatusResponse

logger = structlog.get_logger(__name__)


class DoctorImportService:
    def __init__(self, db: AsyncSession) -> None:
        self.db = db

    async def start_import(self, file_bytes: bytes, redis: Any) -> str:
        task_id = str(uuid4())

        await redis.set(
            f"import:{task_id}",
            json.dumps({"status": "processing", "total_rows": 0, "imported": 0, "errors": []}),
            ex=3600,
        )

        errors: list[dict[str, Any]] = []
        imported = 0
        total_rows = 0

        try:
            wb = load_workbook(filename=BytesIO(file_bytes), read_only=True)
            ws = wb.active
            if ws is None:
                await redis.set(
                    f"import:{task_id}",
                    json.dumps({"status": "error", "total_rows": 0, "imported": 0, "errors": [{"row": 0, "error": "Empty workbook"}]}),
                    ex=3600,
                )
                return task_id

            rows = list(ws.iter_rows(min_row=2, values_only=True))
            total_rows = len(rows)

            for idx, row in enumerate(rows, start=2):
                try:
                    email = str(row[0]).strip().lower() if row[0] else ""
                    first_name = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                    last_name = str(row[2]).strip() if len(row) > 2 and row[2] else ""
                    phone = str(row[3]).strip() if len(row) > 3 and row[3] else ""

                    if not email:
                        errors.append({"row": idx, "error": "Email is required"})
                        continue

                    existing = await self.db.execute(
                        select(User.id).where(func.lower(User.email) == email)
                    )
                    if existing.scalar_one_or_none():
                        errors.append({"row": idx, "error": f"Duplicate email: {email}"})
                        continue

                    temp_password = f"Tmp{uuid4().hex[:12]}!"
                    user = User(
                        email=email,
                        password_hash=hash_password(temp_password),
                        email_verified_at=datetime.now(UTC),
                        is_active=True,
                    )
                    self.db.add(user)
                    await self.db.flush()

                    imp_slug = await generate_unique_slug(
                        self.db,
                        DoctorProfile,
                        f"{last_name or 'N-A'} {first_name or 'N-A'}",
                    )
                    profile = DoctorProfile(
                        user_id=user.id,
                        first_name=first_name or "N/A",
                        last_name=last_name or "N/A",
                        phone=phone or "",
                        status=DoctorStatus.APPROVED,
                        slug=imp_slug,
                    )
                    self.db.add(profile)
                    await self.db.flush()
                    imported += 1

                except Exception as exc:
                    errors.append({"row": idx, "error": str(exc)})

            await self.db.commit()
        except Exception as exc:
            logger.error("import_failed", error=str(exc))
            errors.append({"row": 0, "error": f"File processing error: {exc}"})

        result_data = {
            "status": "completed" if not errors else "completed",
            "total_rows": total_rows,
            "imported": imported,
            "errors": errors,
        }
        await redis.set(f"import:{task_id}", json.dumps(result_data), ex=3600)
        return task_id

    async def get_import_status(self, task_id: str, redis: Any) -> ImportStatusResponse:
        raw = await redis.get(f"import:{task_id}")
        if not raw:
            raise NotFoundError("Import task not found")
        data = json.loads(raw)
        return ImportStatusResponse(
            status=data["status"],
            total_rows=data.get("total_rows", 0),
            imported=data.get("imported", 0),
            errors=[ImportErrorItem(**e) for e in data.get("errors", [])],
        )
