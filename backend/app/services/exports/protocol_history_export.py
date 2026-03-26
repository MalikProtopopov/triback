"""XLSX export: protocol history journal (management)."""

from __future__ import annotations

from datetime import date
from typing import Any
from uuid import UUID

from openpyxl import Workbook
from sqlalchemy import false as sql_false
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import aliased

from app.models.cities import City
from app.models.profiles import DoctorProfile
from app.models.protocol_history import ProtocolHistoryEntry
from app.models.users import User
from app.services.exports.doctors_export import _best_subscription_subquery
from app.services.exports.export_errors import ExportTooLargeError
from app.services.exports.limits import MAX_EXPORT_ROWS
from app.services.exports.msk import format_date_msk, format_dt_msk, msk_range_to_utc_exclusive_end
from app.services.exports.translations import ru_doctor_status, ru_protocol_action
from app.services.exports.xlsx_base import (
    apply_autofilter,
    bold_font,
    cell_value,
    workbook_to_bytes,
    write_header_row,
)

PROTOCOL_HEADERS = [
    "ID записи",
    "Год протокола",
    "Название / номер протокола",
    "Тип решения (EN)",
    "Тип решения (RU)",
    "Примечания",
    "Дата создания записи",
    "Дата последнего изменения",
    "ID врача (user_id)",
    "Email врача",
    "Фамилия врача",
    "Имя врача",
    "Отчество врача",
    "Телефон врача",
    "Текущий статус врача (EN)",
    "Текущий статус (RU)",
    "Создал: email",
    "Создал: ФИО",
    "Последний редактор: email",
]

ACTIVE_DOCTOR_HEADERS = [
    "ID пользователя",
    "Email (учётный)",
    "Фамилия",
    "Имя",
    "Отчество",
    "Телефон",
    "Город",
    "Статус профиля (RU)",
    "Название плана",
    "Дата окончания подписки",
]


def _fio(first: str | None, last: str | None, middle: str | None) -> str:
    parts = [last or "", first or "", middle or ""]
    return " ".join(p for p in parts if p).strip()


async def _active_doctor_rows(
    db: AsyncSession,
) -> list[tuple[Any, ...]]:
    """Врачи: status=active, не удалён, не исключён из членства."""
    best_sub = _best_subscription_subquery()
    q = (
        select(
            User.id,
            User.email,
            DoctorProfile.last_name,
            DoctorProfile.first_name,
            DoctorProfile.middle_name,
            DoctorProfile.phone,
            City.name,
            DoctorProfile.status,
            best_sub.c.plan_name,
            best_sub.c.sub_ends,
        )
        .select_from(
            DoctorProfile.__table__.join(User.__table__, DoctorProfile.user_id == User.id)
            .outerjoin(City.__table__, DoctorProfile.city_id == City.id)
            .outerjoin(best_sub, best_sub.c.sub_uid == User.id)
        )
        .where(
            DoctorProfile.status == "active",
            DoctorProfile.is_deleted.is_(False),
            DoctorProfile.membership_excluded_at.is_(None),
        )
        .order_by(DoctorProfile.last_name.asc(), DoctorProfile.first_name.asc())
    )
    return list((await db.execute(q)).all())


def _protocol_base_joins():
    DocUser = aliased(User)
    CreUser = aliased(User)
    EdUser = aliased(User)
    DocDp = aliased(DoctorProfile)
    CreDp = aliased(DoctorProfile)
    return DocUser, CreUser, EdUser, DocDp, CreDp


def _protocol_where(
    *,
    date_from: date | None,
    date_to: date | None,
    apply_date_filter: bool,
    years: list[int] | None,
    action_types: list[str] | None,
    doctor_user_id: UUID | None,
    created_by_user_id: UUID | None,
    doctor_ids_in: list[UUID] | None,
) -> list[Any]:
    e = ProtocolHistoryEntry
    conds: list[Any] = []
    if apply_date_filter and date_from is not None and date_to is not None:
        lo, hi = msk_range_to_utc_exclusive_end(date_from, date_to)
        conds.append(e.created_at >= lo)
        conds.append(e.created_at < hi)
    if years:
        conds.append(e.year.in_(years))
    if action_types:
        conds.append(e.action_type.in_(action_types))
    if doctor_user_id:
        conds.append(e.doctor_user_id == doctor_user_id)
    if created_by_user_id:
        conds.append(e.created_by_user_id == created_by_user_id)
    if doctor_ids_in is not None:
        if len(doctor_ids_in) == 0:
            conds.append(sql_false())
        else:
            conds.append(e.doctor_user_id.in_(doctor_ids_in))
    return conds


async def build_protocol_history_xlsx(
    db: AsyncSession,
    *,
    date_from: date | None = None,
    date_to: date | None = None,
    apply_date_filter: bool = True,
    years: list[int] | None = None,
    action_types: list[str] | None = None,
    doctor_user_id: UUID | None = None,
    created_by_user_id: UUID | None = None,
    active_doctors_only: bool = False,
) -> bytes:
    DocUser, CreUser, EdUser, DocDp, CreDp = _protocol_base_joins()
    entry = ProtocolHistoryEntry

    doctor_ids_in: list[UUID] | None = None
    active_sheet_rows: list[tuple[Any, ...]] | None = None
    if active_doctors_only:
        active_sheet_rows = await _active_doctor_rows(db)
        doctor_ids_in = [r[0] for r in active_sheet_rows]

    conds = _protocol_where(
        date_from=date_from,
        date_to=date_to,
        apply_date_filter=apply_date_filter,
        years=years,
        action_types=action_types,
        doctor_user_id=doctor_user_id,
        created_by_user_id=created_by_user_id,
        doctor_ids_in=doctor_ids_in,
    )

    base_from = (
        entry.__table__.join(DocUser, entry.doctor_user_id == DocUser.id)
        .outerjoin(DocDp, DocDp.user_id == DocUser.id)
        .join(CreUser, entry.created_by_user_id == CreUser.id)
        .outerjoin(CreDp, CreDp.user_id == CreUser.id)
        .outerjoin(EdUser, entry.last_edited_by_user_id == EdUser.id)
    )

    cnt_q = select(func.count(entry.id)).select_from(base_from)
    if conds:
        cnt_q = cnt_q.where(*conds)
    total = (await db.execute(cnt_q)).scalar() or 0
    if total > MAX_EXPORT_ROWS:
        raise ExportTooLargeError(total)

    q = (
        select(
            entry,
            DocUser.email.label("doc_email"),
            DocDp.last_name.label("doc_ln"),
            DocDp.first_name.label("doc_fn"),
            DocDp.middle_name.label("doc_mn"),
            DocDp.phone.label("doc_phone"),
            DocDp.status.label("doc_status"),
            CreUser.email.label("cre_email"),
            CreDp.last_name.label("cre_ln"),
            CreDp.first_name.label("cre_fn"),
            CreDp.middle_name.label("cre_mn"),
            EdUser.email.label("ed_email"),
        )
        .select_from(base_from)
        .order_by(entry.year.desc(), entry.created_at.desc())
    )
    if conds:
        q = q.where(*conds)

    prot_rows = (await db.execute(q)).all()

    admission_n = sum(1 for r in prot_rows if r[0].action_type == "admission")
    exclusion_n = sum(1 for r in prot_rows if r[0].action_type == "exclusion")

    wb = Workbook()
    if active_doctors_only and active_sheet_rows is not None:
        ws0 = wb.active
        ws0.title = "Врачи"
        write_header_row(ws0, 1, ACTIVE_DOCTOR_HEADERS)
        for ri, r in enumerate(active_sheet_rows, start=2):
            uid, email, ln, fn, mn, phone, city, st, plan_name, sub_ends = r
            plan = plan_name or ""
            ends = format_date_msk(sub_ends) if sub_ends else ""
            vals = [
                str(uid),
                email,
                ln,
                fn,
                mn,
                phone,
                city,
                ru_doctor_status(str(st)),
                plan,
                ends,
            ]
            for ci, v in enumerate(vals, start=1):
                ws0.cell(row=ri, column=ci, value=cell_value(v))
        last_ad = 1 + len(active_sheet_rows)
        apply_autofilter(ws0, len(ACTIVE_DOCTOR_HEADERS), max(1, last_ad))
        ws = wb.create_sheet("Протоколы")
    else:
        ws = wb.active
        ws.title = "Протоколы"

    write_header_row(ws, 1, PROTOCOL_HEADERS)
    for ri, row in enumerate(prot_rows, start=2):
        e: ProtocolHistoryEntry = row[0]
        doc_email = row[1]
        doc_ln, doc_fn, doc_mn = row[2], row[3], row[4]
        doc_phone = row[5]
        doc_status = row[6]
        cre_email = row[7]
        cre_ln, cre_fn, cre_mn = row[8], row[9], row[10]
        ed_email = row[11]

        cre_fio = _fio(cre_fn, cre_ln, cre_mn) if cre_ln or cre_fn else ""

        vals = [
            str(e.id),
            e.year,
            e.protocol_title,
            e.action_type,
            ru_protocol_action(e.action_type),
            e.notes,
            format_dt_msk(e.created_at),
            format_dt_msk(e.updated_at),
            str(e.doctor_user_id),
            doc_email,
            doc_ln,
            doc_fn,
            doc_mn,
            doc_phone,
            str(doc_status) if doc_status else "",
            ru_doctor_status(str(doc_status)) if doc_status else "",
            cre_email,
            cre_fio,
            ed_email or "",
        ]
        for ci, v in enumerate(vals, start=1):
            ws.cell(row=ri, column=ci, value=cell_value(v))

    last_pr = 1 + len(prot_rows)
    tot = last_pr + 1
    ws.cell(row=tot, column=1, value="Итого")
    ws.cell(row=tot, column=1).font = bold_font()
    ws.cell(row=tot, column=2, value=f"Приёмов: {admission_n}")
    ws.cell(row=tot, column=2).font = bold_font()
    ws.cell(row=tot, column=3, value=f"Исключений: {exclusion_n}")
    ws.cell(row=tot, column=3).font = bold_font()
    apply_autofilter(ws, len(PROTOCOL_HEADERS), tot)

    return workbook_to_bytes(wb)
