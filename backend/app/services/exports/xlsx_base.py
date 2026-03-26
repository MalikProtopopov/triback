"""openpyxl helpers: workbook, autofilter, optional totals."""

from __future__ import annotations

import io
from decimal import Decimal
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook as WorkbookType


def new_workbook(sheet_title: str) -> tuple[WorkbookType, Any]:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]
    return wb, ws


def write_header_row(ws, col: int, headers: list[str]) -> None:
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)


def cell_value(v: Any) -> Any:
    """None -> empty cell; Decimal preserved for Excel."""
    if v is None:
        return None
    if isinstance(v, float) and v != v:  # NaN
        return None
    return v


def apply_autofilter(ws, num_cols: int, num_rows: int) -> None:
    if num_rows < 1:
        return
    last_col = get_column_letter(num_cols)
    ws.auto_filter.ref = f"A1:{last_col}{num_rows}"


def workbook_to_bytes(wb: WorkbookType) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def bold_font() -> Font:
    return Font(bold=True)


def set_totals_row(
    ws,
    row: int,
    *,
    label_col: int,
    label: str,
    sum_col: int,
    total: Decimal | float | int,
) -> None:
    c = ws.cell(row=row, column=label_col, value=label)
    c.font = bold_font()
    v = ws.cell(row=row, column=sum_col, value=float(total))
    v.font = bold_font()
