"""Admin doctor management service — list, detail, moderation, drafts, import."""

from __future__ import annotations

import json
from datetime import UTC, datetime, timedelta
from io import BytesIO
from typing import Any
from uuid import UUID, uuid4

import structlog
from openpyxl import load_workbook
from sqlalchemy import Select, and_, exists, func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import joinedload, selectinload

from app.core.config import settings
from app.core.exceptions import ConflictError, NotFoundError
from app.core.security import hash_password
from app.core.utils import generate_unique_slug
from app.models.profiles import (
    DoctorProfile,
    DoctorProfileChange,
    DoctorSpecialization,
    ModerationHistory,
)
from app.models.subscriptions import Payment, Subscription
from app.models.users import Role, User, UserRoleAssignment
from app.schemas.doctor_admin import (
    AdminCreateDoctorResponse,
    CityNested,
    DoctorDetailResponse,
    DoctorListItemResponse,
    DocumentNested,
    ImportErrorItem,
    ImportStatusResponse,
    ModerationHistoryNested,
    PaymentNested,
    PendingDraftNested,
    PortalUserDetailResponse,
    PortalUserListItem,
    SubscriptionNested,
)
from app.tasks.email_tasks import (
    send_custom_email,
    send_doctor_invite_email,
    send_draft_result_notification,
    send_moderation_result_notification,
    send_reminder_notification,
)

logger = structlog.get_logger(__name__)


class DoctorAdminService:
    def __init__(self, db: AsyncSession) -> None:
        self.db = db

    # ── helpers ───────────────────────────────────────────────────

    async def _get_profile_or_404(self, profile_id: UUID) -> DoctorProfile:
        result = await self.db.execute(
            select(DoctorProfile).where(DoctorProfile.id == profile_id)
        )
        profile = result.scalar_one_or_none()
        if not profile:
            raise NotFoundError("Doctor profile not found")
        return profile

    # ── 0. create doctor (manual) ─────────────────────────────────

    async def create_doctor(self, data: dict[str, Any]) -> AdminCreateDoctorResponse:
        existing = await self.db.execute(
            select(User.id).where(User.email == data["email"])
        )
        if existing.scalar_one_or_none():
            raise ConflictError("User with this email already exists")

        doctor_role = (
            await self.db.execute(select(Role).where(Role.name == "doctor"))
        ).scalar_one_or_none()
        if not doctor_role:
            raise NotFoundError("Role 'doctor' not found in database")

        temp_password = f"Tmp{uuid4().hex[:12]}!"

        user = User(
            email=data["email"],
            password_hash=hash_password(temp_password),
            email_verified_at=datetime.now(UTC),
            is_active=True,
        )
        self.db.add(user)
        await self.db.flush()

        self.db.add(UserRoleAssignment(user_id=user.id, role_id=doctor_role.id))

        slug = await generate_unique_slug(
            self.db, DoctorProfile, f"{data['last_name']} {data['first_name']}"
        )

        profile = DoctorProfile(
            user_id=user.id,
            first_name=data["first_name"],
            last_name=data["last_name"],
            phone=data["phone"],
            middle_name=data.get("middle_name"),
            city_id=data.get("city_id"),
            clinic_name=data.get("clinic_name"),
            position=data.get("position"),
            academic_degree=data.get("academic_degree"),
            bio=data.get("bio"),
            public_email=data.get("public_email"),
            public_phone=data.get("public_phone"),
            status=data.get("status", "approved"),
            slug=slug,
        )
        self.db.add(profile)
        await self.db.flush()

        spec_ids = data.get("specialization_ids") or []
        for sid in spec_ids:
            self.db.add(
                DoctorSpecialization(
                    doctor_profile_id=profile.id,
                    specialization_id=sid,
                )
            )

        await self.db.commit()

        if data.get("send_invite", True):
            await send_doctor_invite_email.kiq(
                data["email"], temp_password, settings.FRONTEND_URL
            )

        logger.info(
            "doctor_created_by_admin",
            user_id=str(user.id),
            profile_id=str(profile.id),
            email=data["email"],
        )

        return AdminCreateDoctorResponse(
            user_id=user.id,
            profile_id=profile.id,
            email=user.email,
            first_name=profile.first_name,
            last_name=profile.last_name,
            status=profile.status,
            temp_password=temp_password if not data.get("send_invite", True) else None,
        )

    # ── 1. list doctors ───────────────────────────────────────────

    async def list_doctors(
        self,
        *,
        limit: int = 20,
        offset: int = 0,
        status: str | None = None,
        subscription_status: str | None = None,
        city_id: UUID | None = None,
        has_data_changed: bool | None = None,
        search: str | None = None,
        sort_by: str = "created_at",
        sort_order: str = "desc",
    ) -> dict[str, Any]:
        base = (
            select(DoctorProfile)
            .join(User, DoctorProfile.user_id == User.id)
            .options(joinedload(DoctorProfile.user), joinedload(DoctorProfile.city))
        )

        count_q = (
            select(func.count(DoctorProfile.id))
            .join(User, DoctorProfile.user_id == User.id)
        )

        filters: list[Any] = []

        if status:
            filters.append(DoctorProfile.status == status)
        if city_id:
            filters.append(DoctorProfile.city_id == city_id)
        if search and len(search) >= 2:
            pattern = f"%{search}%"
            filters.append(
                or_(
                    DoctorProfile.last_name.ilike(pattern),
                    DoctorProfile.first_name.ilike(pattern),
                    User.email.ilike(pattern),
                )
            )
        if has_data_changed is True:
            pending_exists = exists(
                select(DoctorProfileChange.id).where(
                    and_(
                        DoctorProfileChange.doctor_profile_id == DoctorProfile.id,
                        DoctorProfileChange.status == "pending",
                    )
                )
            )
            filters.append(pending_exists)

        need_sub_join = subscription_status is not None or sort_by == "subscription_ends_at"

        if need_sub_join:
            latest_sub = (
                select(
                    Subscription.user_id,
                    Subscription.status.label("sub_status"),
                    Subscription.ends_at.label("sub_ends_at"),
                    func.row_number()
                    .over(partition_by=Subscription.user_id, order_by=Subscription.created_at.desc())
                    .label("rn"),
                )
                .subquery("latest_sub")
            )
            sub_sq = (
                select(
                    latest_sub.c.user_id,
                    latest_sub.c.sub_status,
                    latest_sub.c.sub_ends_at,
                )
                .where(latest_sub.c.rn == 1)
                .subquery("sub_sq")
            )
            base = base.outerjoin(sub_sq, DoctorProfile.user_id == sub_sq.c.user_id)
            count_q = count_q.outerjoin(sub_sq, DoctorProfile.user_id == sub_sq.c.user_id)

            if subscription_status == "expiring_soon":
                now = datetime.now(UTC)
                filters.append(
                    and_(
                        sub_sq.c.sub_status == "active",
                        sub_sq.c.sub_ends_at.between(now, now + timedelta(days=7)),
                    )
                )
            elif subscription_status == "never":
                filters.append(sub_sq.c.sub_status.is_(None))
            elif subscription_status:
                filters.append(sub_sq.c.sub_status == subscription_status)

        if filters:
            base = base.where(and_(*filters))
            count_q = count_q.where(and_(*filters))

        total = (await self.db.execute(count_q)).scalar() or 0

        base = self._apply_sort(base, sort_by, sort_order, need_sub_join)
        base = base.offset(offset).limit(limit)

        rows = (await self.db.execute(base)).unique().scalars().all()

        user_ids = [dp.user_id for dp in rows]
        profile_ids = [dp.id for dp in rows]

        sub_map: dict[UUID, SubscriptionNested] = {}
        if user_ids:
            sub_q = (
                select(Subscription)
                .options(joinedload(Subscription.plan))
                .where(Subscription.user_id.in_(user_ids))
                .order_by(Subscription.user_id, Subscription.created_at.desc())
            )
            sub_rows = (await self.db.execute(sub_q)).unique().scalars().all()
            for s in sub_rows:
                if s.user_id not in sub_map:
                    sub_map[s.user_id] = SubscriptionNested(
                        id=s.id, status=s.status,
                        plan_name=s.plan.name if s.plan else None,
                        starts_at=s.starts_at, ends_at=s.ends_at,
                    )

        pending_set: set[UUID] = set()
        if profile_ids:
            pending_q = (
                select(DoctorProfileChange.doctor_profile_id)
                .where(
                    DoctorProfileChange.doctor_profile_id.in_(profile_ids),
                    DoctorProfileChange.status == "pending",
                )
                .distinct()
            )
            pending_set = set((await self.db.execute(pending_q)).scalars().all())

        items: list[DoctorListItemResponse] = []
        for dp in rows:
            items.append(
                DoctorListItemResponse(
                    id=dp.id,
                    user_id=dp.user_id,
                    email=dp.user.email,
                    first_name=dp.first_name,
                    last_name=dp.last_name,
                    middle_name=dp.middle_name,
                    phone=dp.phone,
                    city=CityNested(id=dp.city.id, name=dp.city.name) if dp.city else None,
                    specialization=dp.specialization.name if dp.specialization else None,
                    moderation_status=dp.status,
                    has_medical_diploma=dp.has_medical_diploma,
                    subscription=sub_map.get(dp.user_id),
                    has_pending_changes=dp.id in pending_set,
                    created_at=dp.created_at,
                )
            )

        return {"data": items, "total": total, "limit": limit, "offset": offset}

    def _apply_sort(
        self, q: Select, sort_by: str, sort_order: str, has_sub_join: bool  # type: ignore[type-arg]
    ) -> Select:  # type: ignore[type-arg]
        col: Any
        if sort_by == "last_name":
            col = DoctorProfile.last_name
        elif sort_by == "subscription_ends_at" and has_sub_join:
            col = func.coalesce(
                select(Subscription.ends_at)
                .where(Subscription.user_id == DoctorProfile.user_id)
                .order_by(Subscription.created_at.desc())
                .limit(1)
                .correlate(DoctorProfile)
                .scalar_subquery(),
                datetime.min,
            )
        else:
            col = DoctorProfile.created_at

        return q.order_by(col.desc() if sort_order == "desc" else col.asc())

    async def _latest_subscription_nested(self, user_id: UUID) -> SubscriptionNested | None:
        result = await self.db.execute(
            select(Subscription)
            .options(joinedload(Subscription.plan))
            .where(Subscription.user_id == user_id)
            .order_by(Subscription.created_at.desc())
            .limit(1)
        )
        sub = result.scalar_one_or_none()
        if not sub:
            return None
        return SubscriptionNested(
            id=sub.id,
            status=sub.status,
            plan_name=sub.plan.name if sub.plan else None,
            starts_at=sub.starts_at,
            ends_at=sub.ends_at,
        )

    async def _has_pending_changes(self, profile_id: UUID) -> bool:
        result = await self.db.execute(
            select(func.count(DoctorProfileChange.id)).where(
                and_(
                    DoctorProfileChange.doctor_profile_id == profile_id,
                    DoctorProfileChange.status == "pending",
                )
            )
        )
        return (result.scalar() or 0) > 0

    # ── 2. get doctor detail ──────────────────────────────────────

    async def get_doctor(self, profile_id: UUID) -> DoctorDetailResponse:
        result = await self.db.execute(
            select(DoctorProfile)
            .options(
                joinedload(DoctorProfile.user),
                joinedload(DoctorProfile.city),
                joinedload(DoctorProfile.specialization),
                selectinload(DoctorProfile.documents),
                selectinload(DoctorProfile.profile_changes),
            )
            .where(DoctorProfile.id == profile_id)
        )
        dp = result.unique().scalar_one_or_none()
        if not dp:
            raise NotFoundError("Doctor profile not found")

        sub_info = await self._latest_subscription_nested(dp.user_id)

        payments_result = await self.db.execute(
            select(Payment)
            .where(Payment.user_id == dp.user_id)
            .order_by(Payment.created_at.desc())
            .limit(20)
        )
        payments_rows = payments_result.scalars().all()

        pending_draft: PendingDraftNested | None = None
        for pc in dp.profile_changes:
            if pc.status == "pending":
                pending_draft = PendingDraftNested(
                    id=pc.id,
                    changes=pc.changes,
                    changed_fields=pc.changed_fields,
                    status=pc.status,
                    moderation_comment=pc.moderation_comment,
                    submitted_at=pc.submitted_at,
                    rejection_reason=pc.rejection_reason,
                )
                break

        mod_history_result = await self.db.execute(
            select(ModerationHistory)
            .where(
                and_(
                    ModerationHistory.entity_type == "doctor_profile",
                    ModerationHistory.entity_id == profile_id,
                )
            )
            .order_by(ModerationHistory.created_at.desc())
        )
        mod_rows = mod_history_result.scalars().all()

        admin_ids = {mh.admin_id for mh in mod_rows if mh.admin_id}
        admin_email_map: dict[UUID, str] = {}
        if admin_ids:
            admin_q = await self.db.execute(
                select(User.id, User.email).where(User.id.in_(admin_ids))
            )
            for row in admin_q.all():
                admin_email_map[row.id] = row.email

        mod_items: list[ModerationHistoryNested] = []
        for mh in mod_rows:
            mod_items.append(
                ModerationHistoryNested(
                    id=mh.id,
                    admin_email=admin_email_map.get(mh.admin_id),
                    action=mh.action,
                    comment=mh.comment,
                    created_at=mh.created_at,
                )
            )

        return DoctorDetailResponse(
            id=dp.id,
            user_id=dp.user_id,
            email=dp.user.email,
            first_name=dp.first_name,
            last_name=dp.last_name,
            middle_name=dp.middle_name,
            phone=dp.phone,
            passport_data=dp.passport_data,
            city=CityNested(id=dp.city.id, name=dp.city.name) if dp.city else None,
            clinic_name=dp.clinic_name,
            position=dp.position,
            specialization=dp.specialization.name if dp.specialization else None,
            academic_degree=dp.academic_degree,
            bio=dp.bio,
            public_email=dp.public_email,
            public_phone=dp.public_phone,
            photo_url=dp.photo_url,
            moderation_status=dp.status,
            has_medical_diploma=dp.has_medical_diploma,
            diploma_photo_url=dp.diploma_photo_url,
            slug=dp.slug,
            documents=[
                DocumentNested(
                    id=d.id,
                    document_type=d.document_type,
                    original_filename=d.original_filename,
                    file_url=d.file_url,
                    file_size=d.file_size,
                    mime_type=d.mime_type,
                    uploaded_at=d.uploaded_at,
                )
                for d in dp.documents
            ],
            subscription=sub_info,
            payments=[
                PaymentNested(
                    id=p.id,
                    amount=float(p.amount),
                    product_type=p.product_type,
                    status=p.status,
                    paid_at=p.paid_at,
                    created_at=p.created_at,
                )
                for p in payments_rows
            ],
            pending_draft=pending_draft,
            moderation_history=mod_items,
            created_at=dp.created_at,
        )

    # ── 3. moderate ───────────────────────────────────────────────

    async def moderate(
        self,
        profile_id: UUID,
        admin_id: UUID,
        action: str,
        comment: str | None = None,
    ) -> str:
        dp = await self._get_profile_or_404(profile_id)

        new_status = "approved" if action == "approve" else "rejected"
        dp.status = new_status  # type: ignore[assignment]

        if action == "approve" and not dp.slug:
            dp.slug = await generate_unique_slug(
                self.db, DoctorProfile, f"{dp.last_name} {dp.first_name}"
            )

        self.db.add(
            ModerationHistory(
                admin_id=admin_id,
                entity_type="doctor_profile",
                entity_id=profile_id,
                action=action,
                comment=comment,
            )
        )

        await self.db.flush()

        user = await self.db.get(User, dp.user_id)
        if user:
            await send_moderation_result_notification.kiq(user.email, new_status, comment)

        await self.db.commit()
        return new_status

    # ── 4. approve draft ──────────────────────────────────────────

    async def approve_draft(
        self,
        profile_id: UUID,
        admin_id: UUID,
        action: str,
        rejection_reason: str | None = None,
    ) -> str:
        dp = await self._get_profile_or_404(profile_id)

        result = await self.db.execute(
            select(DoctorProfileChange).where(
                and_(
                    DoctorProfileChange.doctor_profile_id == profile_id,
                    DoctorProfileChange.status == "pending",
                )
            )
        )
        draft = result.scalar_one_or_none()
        if not draft:
            raise NotFoundError("No pending draft found")

        now = datetime.now(UTC)

        if action == "approve":
            for key, value in draft.changes.items():
                if hasattr(dp, key):
                    setattr(dp, key, value)
            draft.status = "approved"  # type: ignore[assignment]
            draft.reviewed_at = now
            draft.reviewed_by = admin_id
            msg = "Changes approved and applied"
        else:
            draft.status = "rejected"  # type: ignore[assignment]
            draft.reviewed_at = now
            draft.reviewed_by = admin_id
            draft.rejection_reason = rejection_reason
            msg = "Changes rejected"

        self.db.add(
            ModerationHistory(
                admin_id=admin_id,
                entity_type="doctor_profile_change",
                entity_id=draft.id,
                action=f"draft_{action}",
                comment=rejection_reason if action == "reject" else None,
            )
        )
        await self.db.flush()

        user = await self.db.get(User, dp.user_id)
        if user:
            await send_draft_result_notification.kiq(
                user.email, action, rejection_reason
            )

        await self.db.commit()
        return msg

    # ── 5. toggle active ─────────────────────────────────────────

    async def toggle_active(self, profile_id: UUID, admin_id: UUID, is_public: bool) -> bool:
        dp = await self._get_profile_or_404(profile_id)

        if is_public:
            dp.status = "active"  # type: ignore[assignment]
        else:
            dp.status = "deactivated"  # type: ignore[assignment]

        self.db.add(
            ModerationHistory(
                admin_id=admin_id,
                entity_type="doctor_profile",
                entity_id=profile_id,
                action="activate" if is_public else "deactivate",
            )
        )
        await self.db.commit()
        return is_public

    # ── 6. send reminder ──────────────────────────────────────────

    async def send_reminder(
        self, profile_id: UUID, message: str | None = None
    ) -> None:
        dp = await self._get_profile_or_404(profile_id)
        user = await self.db.get(User, dp.user_id)
        if not user:
            raise NotFoundError("User not found")
        await send_reminder_notification.kiq(user.email, message)

    # ── 6b. send email ────────────────────────────────────────────

    async def send_email(
        self, profile_id: UUID, subject: str, body: str
    ) -> None:
        dp = await self._get_profile_or_404(profile_id)
        user = await self.db.get(User, dp.user_id)
        if not user:
            raise NotFoundError("User not found")
        await send_custom_email.kiq(user.email, subject, body)

    # ── 7. import from excel ──────────────────────────────────────

    async def start_import(self, file_bytes: bytes, redis: Any) -> str:
        """Parse Excel, create doctors row by row, store result in Redis."""
        task_id = str(uuid4())

        await redis.set(
            f"import:{task_id}",
            json.dumps({"status": "processing", "total_rows": 0, "imported": 0, "errors": []}),
            ex=3600,
        )

        errors: list[dict[str, Any]] = []
        imported = 0
        total_rows = 0

        try:
            wb = load_workbook(filename=BytesIO(file_bytes), read_only=True)
            ws = wb.active
            if ws is None:
                await redis.set(
                    f"import:{task_id}",
                    json.dumps({"status": "error", "total_rows": 0, "imported": 0, "errors": [{"row": 0, "error": "Empty workbook"}]}),
                    ex=3600,
                )
                return task_id

            rows = list(ws.iter_rows(min_row=2, values_only=True))
            total_rows = len(rows)

            for idx, row in enumerate(rows, start=2):
                try:
                    email = str(row[0]).strip() if row[0] else ""
                    first_name = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                    last_name = str(row[2]).strip() if len(row) > 2 and row[2] else ""
                    phone = str(row[3]).strip() if len(row) > 3 and row[3] else ""

                    if not email:
                        errors.append({"row": idx, "error": "Email is required"})
                        continue

                    existing = await self.db.execute(
                        select(User.id).where(User.email == email)
                    )
                    if existing.scalar_one_or_none():
                        errors.append({"row": idx, "error": f"Duplicate email: {email}"})
                        continue

                    temp_password = f"Tmp{uuid4().hex[:12]}!"
                    user = User(
                        email=email,
                        password_hash=hash_password(temp_password),
                        email_verified_at=datetime.now(UTC),
                        is_active=True,
                    )
                    self.db.add(user)
                    await self.db.flush()

                    imp_slug = await generate_unique_slug(
                        self.db,
                        DoctorProfile,
                        f"{last_name or 'N-A'} {first_name or 'N-A'}",
                    )
                    profile = DoctorProfile(
                        user_id=user.id,
                        first_name=first_name or "N/A",
                        last_name=last_name or "N/A",
                        phone=phone or "",
                        status="approved",
                        slug=imp_slug,
                    )
                    self.db.add(profile)
                    await self.db.flush()
                    imported += 1

                except Exception as exc:
                    errors.append({"row": idx, "error": str(exc)})

            await self.db.commit()
        except Exception as exc:
            logger.error("import_failed", error=str(exc))
            errors.append({"row": 0, "error": f"File processing error: {exc}"})

        result_data = {
            "status": "completed" if not errors else "completed",
            "total_rows": total_rows,
            "imported": imported,
            "errors": errors,
        }
        await redis.set(f"import:{task_id}", json.dumps(result_data), ex=3600)
        return task_id

    async def get_import_status(self, task_id: str, redis: Any) -> ImportStatusResponse:
        raw = await redis.get(f"import:{task_id}")
        if not raw:
            raise NotFoundError("Import task not found")
        data = json.loads(raw)
        return ImportStatusResponse(
            status=data["status"],
            total_rows=data.get("total_rows", 0),
            imported=data.get("imported", 0),
            errors=[ImportErrorItem(**e) for e in data.get("errors", [])],
        )

    # ── 8. portal users ───────────────────────────────────────────

    async def list_portal_users(
        self,
        *,
        limit: int = 20,
        offset: int = 0,
        search: str | None = None,
        sort_by: str = "created_at",
        sort_order: str = "desc",
    ) -> dict[str, Any]:
        admin_roles = {"admin", "manager", "accountant"}

        admin_subq = (
            select(UserRoleAssignment.user_id)
            .join(Role, UserRoleAssignment.role_id == Role.id)
            .where(Role.name.in_(list(admin_roles)))
            .subquery()
        )

        base = select(User).where(~User.id.in_(select(admin_subq)))
        count_q = select(func.count(User.id)).where(~User.id.in_(select(admin_subq)))

        if search and len(search) >= 2:
            pattern = f"%{search}%"
            base = base.where(User.email.ilike(pattern))
            count_q = count_q.where(User.email.ilike(pattern))

        total = (await self.db.execute(count_q)).scalar() or 0

        sort_col = User.created_at
        base = base.order_by(sort_col.desc() if sort_order == "desc" else sort_col.asc())
        base = base.offset(offset).limit(limit)

        users = (await self.db.execute(base)).scalars().all()
        u_ids = [u.id for u in users]

        roles_map: dict[UUID | str, list[str]] = {}
        if u_ids:
            role_q = await self.db.execute(
                select(UserRoleAssignment.user_id, Role.name)
                .join(Role, UserRoleAssignment.role_id == Role.id)
                .where(UserRoleAssignment.user_id.in_(u_ids))
            )
            for uid, rname in role_q.all():
                roles_map.setdefault(uid, []).append(rname)

        dp_map: dict[UUID | str, DoctorProfile] = {}
        if u_ids:
            dp_q = await self.db.execute(
                select(DoctorProfile).where(DoctorProfile.user_id.in_(u_ids))
            )
            for dp in dp_q.scalars().all():
                dp_map[dp.user_id] = dp

        sub_map: dict[UUID, SubscriptionNested] = {}
        doctor_user_ids = [uid for uid, roles in roles_map.items() if "doctor" in roles]
        if doctor_user_ids:
            sub_q = (
                select(Subscription)
                .options(joinedload(Subscription.plan))
                .where(Subscription.user_id.in_(doctor_user_ids))
                .order_by(Subscription.user_id, Subscription.created_at.desc())
            )
            for s in (await self.db.execute(sub_q)).unique().scalars().all():
                if s.user_id not in sub_map:
                    sub_map[s.user_id] = SubscriptionNested(
                        id=s.id, status=s.status,
                        plan_name=s.plan.name if s.plan else None,
                        starts_at=s.starts_at, ends_at=s.ends_at,
                    )

        role_display_map = {"doctor": "Врач", "user": "Пользователь"}
        items: list[PortalUserListItem] = []
        for u in users:
            user_roles = roles_map.get(u.id, [])
            display_role = next((r for r in user_roles if r in ("doctor", "user")), None)

            dp = dp_map.get(u.id)
            dp_id = dp.id if dp else None
            full_name = f"{dp.last_name} {dp.first_name}" if dp else None

            items.append(
                PortalUserListItem(
                    id=u.id,
                    email=u.email,
                    full_name=full_name,
                    role=display_role,
                    role_display=role_display_map.get(display_role, "Без роли") if display_role else "Без роли",
                    doctor_profile_id=dp_id,
                    subscription=sub_map.get(u.id) if display_role == "doctor" else None,
                    created_at=u.created_at,
                )
            )

        return {"data": items, "total": total, "limit": limit, "offset": offset}

    async def get_portal_user(self, user_id: UUID) -> PortalUserDetailResponse:
        user = await self.db.get(User, user_id)
        if not user:
            raise NotFoundError("User not found")

        role_result = await self.db.execute(
            select(Role.name)
            .join(UserRoleAssignment, Role.id == UserRoleAssignment.role_id)
            .where(UserRoleAssignment.user_id == user.id)
        )
        user_roles = list(role_result.scalars().all())
        display_role = next((r for r in user_roles if r in ("doctor", "user")), None)
        role_display_map = {"doctor": "Врач", "user": "Пользователь"}

        dp_id: UUID | None = None
        full_name: str | None = None
        dp_status: str | None = None
        sub_info: SubscriptionNested | None = None
        payments_list: list[PaymentNested] = []

        if display_role == "doctor":
            dp_result = await self.db.execute(
                select(DoctorProfile).where(DoctorProfile.user_id == user.id)
            )
            dp = dp_result.scalar_one_or_none()
            if dp:
                dp_id = dp.id
                full_name = f"{dp.last_name} {dp.first_name}"
                dp_status = dp.status
                sub_info = await self._latest_subscription_nested(user.id)

        pay_result = await self.db.execute(
            select(Payment)
            .where(Payment.user_id == user.id)
            .order_by(Payment.created_at.desc())
            .limit(20)
        )
        for p in pay_result.scalars().all():
            payments_list.append(
                PaymentNested(
                    id=p.id,
                    amount=float(p.amount),
                    product_type=p.product_type,
                    status=p.status,
                    paid_at=p.paid_at,
                    created_at=p.created_at,
                )
            )

        return PortalUserDetailResponse(
            id=user.id,
            email=user.email,
            full_name=full_name,
            role=display_role,
            role_display=role_display_map.get(display_role, "Без роли") if display_role else "Без роли",
            is_verified=user.email_verified_at is not None,
            onboarding_status=None,
            doctor_profile_id=dp_id,
            doctor_profile_status=dp_status,
            subscription=sub_info,
            payments=payments_list,
            created_at=user.created_at,
        )
