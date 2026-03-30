"""XLSX export endpoints (finance + management)."""

from datetime import UTC, datetime
from io import BytesIO
from unittest.mock import patch
from uuid import uuid4

from app.core.config import settings

import pytest
from httpx import AsyncClient
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.subscriptions import Payment
from tests.factories import create_user


@pytest.mark.anyio
async def test_export_payments_requires_auth(client: AsyncClient):
    r = await client.get("/api/v1/exports/payments")
    assert r.status_code == 401


@pytest.mark.anyio
async def test_export_payments_forbidden_for_doctor(
    client: AsyncClient,
    auth_headers_doctor: dict[str, str],
):
    r = await client.get(
        "/api/v1/exports/payments",
        headers=auth_headers_doctor,
    )
    assert r.status_code == 403


@pytest.mark.anyio
async def test_export_payments_ok_accountant(
    client: AsyncClient,
    auth_headers_accountant: dict[str, str],
):
    r = await client.get(
        "/api/v1/exports/payments",
        headers=auth_headers_accountant,
        params={
            "date_from": "2025-01-01",
            "date_to": "2025-12-31",
        },
    )
    assert r.status_code == 200
    assert r.headers.get("content-type") == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment" in r.headers.get("content-disposition", "")
    assert "payments_" in r.headers["content-disposition"]

    wb = load_workbook(BytesIO(r.content))
    assert "Платежи" in wb.sheetnames
    ws = wb["Платежи"]
    assert ws.cell(row=1, column=1).value == "ID платежа (системный)"


@pytest.mark.anyio
async def test_export_payments_too_many_rows(
    client: AsyncClient,
    auth_headers_accountant: dict[str, str],
    db_session: AsyncSession,
):
    u = await create_user(db_session)
    now = datetime.now(UTC)
    for _ in range(3):
        db_session.add(
            Payment(
                user_id=u.id,
                amount=100.0,
                product_type="subscription",
                status="succeeded",
                paid_at=now,
                created_at=now,
            )
        )
    await db_session.commit()

    with patch("app.services.exports.payments_export.MAX_EXPORT_ROWS", 2):
        r = await client.get(
            "/api/v1/exports/payments",
            headers=auth_headers_accountant,
            params={
                "date_from": "2020-01-01",
                "date_to": "2030-12-31",
            },
        )
    assert r.status_code == 400
    body = r.json()
    msg = body.get("detail") or (body.get("error") or {}).get("message", "")
    assert "10 000" in str(msg)


@pytest.mark.anyio
async def test_export_arrears_ok(
    client: AsyncClient,
    auth_headers_accountant: dict[str, str],
):
    r = await client.get(
        "/api/v1/exports/arrears",
        headers=auth_headers_accountant,
    )
    assert r.status_code == 200
    wb = load_workbook(BytesIO(r.content))
    assert "Задолженности" in wb.sheetnames


@pytest.mark.anyio
async def test_export_event_registrations_requires_event_or_dates(
    client: AsyncClient,
    auth_headers_accountant: dict[str, str],
):
    r = await client.get(
        "/api/v1/exports/event-registrations",
        headers=auth_headers_accountant,
    )
    assert r.status_code == 422


@pytest.mark.anyio
async def test_export_subscriptions_ok(
    client: AsyncClient,
    auth_headers_accountant: dict[str, str],
):
    r = await client.get(
        "/api/v1/exports/subscriptions",
        headers=auth_headers_accountant,
    )
    assert r.status_code == 200
    wb = load_workbook(BytesIO(r.content))
    assert "Подписки" in wb.sheetnames


@pytest.mark.anyio
async def test_export_event_registrations_by_event_id(
    client: AsyncClient,
    auth_headers_accountant: dict[str, str],
):
    r = await client.get(
        "/api/v1/exports/event-registrations",
        headers=auth_headers_accountant,
        params={"event_id": str(uuid4())},
    )
    assert r.status_code == 200


@pytest.mark.anyio
async def test_export_doctors_ok_accountant(
    client: AsyncClient,
    auth_headers_accountant: dict[str, str],
):
    r = await client.get(
        "/api/v1/exports/doctors",
        headers=auth_headers_accountant,
    )
    assert r.status_code == 200
    assert "spreadsheetml" in (r.headers.get("content-type") or "")
    wb = load_workbook(BytesIO(r.content))
    assert "Врачи" in wb.sheetnames


@pytest.mark.anyio
async def test_export_doctors_ok_manager(
    client: AsyncClient,
    auth_headers_manager: dict[str, str],
):
    r = await client.get(
        "/api/v1/exports/doctors",
        headers=auth_headers_manager,
    )
    assert r.status_code == 200
    assert "spreadsheetml" in (r.headers.get("content-type") or "")
    wb = load_workbook(BytesIO(r.content))
    assert "Врачи" in wb.sheetnames
    ws = wb["Врачи"]
    assert ws.cell(row=1, column=1).value == "ID пользователя"


@pytest.mark.anyio
async def test_export_protocol_history_ok_manager(
    client: AsyncClient,
    auth_headers_manager: dict[str, str],
):
    r = await client.get(
        "/api/v1/exports/protocol-history",
        headers=auth_headers_manager,
    )
    assert r.status_code == 200
    wb = load_workbook(BytesIO(r.content))
    assert "Протоколы" in wb.sheetnames
    ws = wb["Протоколы"]
    assert ws.cell(row=1, column=1).value == "ID записи"


@pytest.mark.anyio
async def test_export_protocol_history_ok_accountant(
    client: AsyncClient,
    auth_headers_accountant: dict[str, str],
):
    r = await client.get(
        "/api/v1/exports/protocol-history",
        headers=auth_headers_accountant,
    )
    assert r.status_code == 200
    wb = load_workbook(BytesIO(r.content))
    assert "Протоколы" in wb.sheetnames
    ws = wb["Протоколы"]
    assert ws.cell(row=1, column=1).value == "ID записи"


@pytest.mark.anyio
async def test_export_protocol_history_by_doctor_no_dates(
    client: AsyncClient,
    auth_headers_manager: dict[str, str],
):
    r = await client.get(
        "/api/v1/exports/protocol-history",
        headers=auth_headers_manager,
        params={"doctor_user_id": str(uuid4())},
    )
    assert r.status_code == 200
    assert "protocol_history_doctor_" in r.headers.get("content-disposition", "")


@pytest.mark.anyio
async def test_export_protocol_history_active_doctors_two_sheets(
    client: AsyncClient,
    auth_headers_admin: dict[str, str],
):
    r = await client.get(
        "/api/v1/exports/protocol-history",
        headers=auth_headers_admin,
        params={"active_doctors_only": "true"},
    )
    assert r.status_code == 200
    wb = load_workbook(BytesIO(r.content))
    assert "Врачи" in wb.sheetnames
    assert "Протоколы" in wb.sheetnames


@pytest.mark.anyio
async def test_export_payments_telegram_503_when_chat_not_configured(
    client: AsyncClient,
    auth_headers_accountant: dict[str, str],
):
    with (
        patch.object(settings, "TELEGRAM_EXPORTS_CHAT_ID", ""),
        patch.object(settings, "TELEGRAM_CHANNEL_ID", ""),
    ):
        r = await client.post(
            "/api/v1/exports/payments/telegram",
            headers=auth_headers_accountant,
            params={
                "date_from": "2025-01-01",
                "date_to": "2025-01-31",
            },
        )
    assert r.status_code == 503
